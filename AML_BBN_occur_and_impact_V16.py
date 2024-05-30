# -*- coding: utf-8 -*-
"""
Created on Sat Mar  9 10:07:16 2024

@author: Admin
"""

import xml.etree.ElementTree as ET
import math
from datetime import datetime
import numpy as np
from pgmpy.models import BayesianNetwork
from pgmpy.factors.discrete import TabularCPD
from collections import defaultdict
from pgmpy.inference import VariableElimination
import networkx as nx
import matplotlib.pyplot as plt
import random


# Get user input for their choice on a use case
def get_user_choice():
    print("Please select a use case:")
    print("1. Distribution Station")
    print("2. Sorting Station")
    
    while True:
        try:
            choice = int(input("Enter the number of your choice: "))
            if choice in [1, 2]:
                #return choice
                return 2
            else:
                print("Invalid choice. Please enter a valid number.")
        except ValueError:
            print("Invalid input. Please enter a number.")

# Main program
if __name__ == "__main__":
    user_choice = get_user_choice()
    
    if user_choice == 1:
        print("You selected Distribution Station. The risk propagation is as follows")
        amlFile = ET.parse('Distribution_station.aml')
        
    elif user_choice == 2:
        print("You selected Sorting Station. The risk propagation is as follows")
        amlFile = ET.parse('Sorting_station.aml')

# Parse the XML file
#amlFile = ET.parse('Distribution_station.aml')
root = amlFile.getroot()

#Important Tags for getting all
instanceHierarchyTag=".//{http://www.dke.de/CAEX}InstanceHierarchy"
internalElementTag=".//{http://www.dke.de/CAEX}InternalElement"
externalInterfaceTag=".//{http://www.dke.de/CAEX}ExternalInterface"
AttributeTag=".//{http://www.dke.de/CAEX}Attribute"
ValueTag=".//{http://www.dke.de/CAEX}Value"
internalLinkTag=".//{http://www.dke.de/CAEX}InternalLink"
#Important Tags over

#Program for User input of time of installation
def get_valid_date():
    while True:
        date_input = input("Please enter a date of first installation and use for the machine (in the format YYYY-MM-DD): ")
        if not date_input:
            return "2024-01-01"  # Default date if no date is given
        try:
            datetime.strptime(date_input, "%Y-%m-%d")
            return date_input
        except ValueError:
            print("Invalid date format. Please try again.")
            

'''def calculate_days_and_hours(start_date):
    # Convert the input string to a datetime object
    start_date = datetime.strptime(start_date, "%Y-%m-%d")

    # Get the current date
    current_date = datetime.now()

    # Calculate the difference between the start date and current date
    time_difference = current_date - start_date

    # Calculate the number of days and remaining seconds
    days = time_difference.days
    remaining_seconds = time_difference.seconds

    # Calculate the remaining hours from the remaining seconds
    remaining_hours = remaining_seconds // 3600

    return days, remaining_hours
'''
def calculate_days_and_hours(start_date):
    # Convert the input string to a datetime object
    start_date = datetime.strptime(start_date, "%Y-%m-%d")

    reference_date_input = input("Enter the reference date (YYYY-MM-DD) or leave blank for the current date: ")
    
    if reference_date_input.strip():  # If user provided a reference date
        reference_date = datetime.strptime(reference_date_input, "%Y-%m-%d")
    else:
        reference_date = datetime.now()

    
    # Calculate the difference between the start date and reference date
    time_difference = reference_date - start_date

    # Calculate the number of days and remaining seconds
    days = time_difference.days
    remaining_seconds = time_difference.seconds

    # Calculate the remaining hours from the remaining seconds
    remaining_hours = remaining_seconds // 3600

    return days, remaining_hours

start_date_str = get_valid_date()
print("You entered:", start_date_str)

days, hours = calculate_days_and_hours(start_date_str)

t=days*24 + (24-hours)
print("Time in hours since installation: ", t, ", Number of days:", days, ", Number of hours:", hours)
#Program for User input of time over

# complete aml model attributes extraction
allinone_attrib = []  
allinone_tags = []
allinone_text = []

name_id_tag_list = []  # name, id, and tag of the component and external interface involved
name_list = []
id_list = []
tag_list = []
RefPartnerBegin_list = []
RefPartnerTerminate_list = []
InternalLinks = []

for k in root.findall('.//'):  # prints everything with value
    allinone_attrib.append(k.attrib)
    allinone_tags.append(k.tag)
    allinone_text.append(k.text)

for i, component_attrib in enumerate(allinone_attrib):
    name = component_attrib.get('Name')
    ID = component_attrib.get('ID')
    RPA = component_attrib.get('RefPartnerSideA')
    RPB = component_attrib.get('RefPartnerSideB')
    
    if name:
        name_list.append(name)
    if ID:
        id_list.append(ID)
    if name and ID:
        tag = allinone_tags[i]
        tag_list.append(tag)
        name_id_tag_list.append({'Name': name, 'ID': ID, 'Tag': tag})
    if RPA:
        RefPartnerBegin_list.append(RPA)
    if RPB:
        RefPartnerTerminate_list.append(RPB)
    if RPA and RPB:
        #InternalLinks.append({RPA, RPB, tag})
        InternalLinks.append({RPA, RPB})

#Dealing with attributes from AML file. I need Failure rate and date of initial use.

probability_data = []
AutomationPyramidStatus = {
    'AssetOfICS/Hardware/Machine': [],
    'AssetOfICS/Hardware/Process device/Actuator': [],
    'AssetOfICS/Hardware/Process device/Sensor': [],
    'AssetOfICS/Hardware/Process device/Controller': [],
    'AssetOfICS/Hardware/Process device/Work station': [],
    'AssetOfICS/User': []
}
HazardinSystem = []
VulnerabilityinSystem = []

internal_elements = root.findall(internalElementTag)

def get_attribute_value(internal_element, attribute_name):
    attribute_tag = internal_element.find(f".//{{http://www.dke.de/CAEX}}Attribute[@Name='{attribute_name}']")
    if attribute_tag is not None:
        value_element = attribute_tag.find(ValueTag)
        if value_element is not None:
            return float(value_element.text)
    return None

def calculate_probability_of_failure(failure_rate_value, t):
    failure_rate = float(failure_rate_value)
    return 1 - math.exp(-(failure_rate * t))

def calculate_probability_of_human_error(human_error_percentage_value, t):
    human_error_in_percent = float(human_error_percentage_value)
    human_error_rate = human_error_in_percent / (100 * 8760)
    return 1 - math.exp(-(human_error_rate * t))

def generate_cpd_values_hazard(num_parents):
    cpd_values = [[0] * (2 ** num_parents) for _ in range(2)]

    for i in range(2 ** num_parents):
        num_ones = bin(i).count('1')
        cpd_values[0][i] = (num_parents - num_ones) / num_parents
        cpd_values[1][i] = 1 - cpd_values[0][i]

    return cpd_values

def generate_cpd_values_impact(num_parents):
    cpd_values = [[0] * (2 ** num_parents) for _ in range(2)]
    current_entry = next((entry for entry in result_list if entry['Element'] == node), None)
    for i in range(2 ** num_parents):
        num_ones = bin(i).count('1')
        cpd_values[0][i] = (num_parents - num_ones) / num_parents
        cpd_values[1][i] = 1 - cpd_values[0][i]
        if num_parents - num_ones == 0:
            cpd_values[0][i] = current_entry['Number of children']
            cpd_values[1][i] = 1 - cpd_values[0][i]
    return cpd_values


# Iterate over InternalElement elements
for internal_element in internal_elements:
    internal_element_id = internal_element.get('ID')
    internal_element_name = internal_element.get('Name')
    ref_base_system_unit_path = internal_element.get('RefBaseSystemUnitPath')
    
    failure_rate_value = get_attribute_value(internal_element, 'FailureRatePerHour')
    probability_of_failure = None
    if failure_rate_value is not None:
        probability_of_failure = calculate_probability_of_failure(failure_rate_value, t)
    
    probability_of_exposure_value = get_attribute_value(internal_element, 'Probability of exposure')
    probability_of_exposure = None
    if probability_of_exposure_value is not None:
        probability_of_exposure = probability_of_exposure_value
    
    probability_of_impact_value = get_attribute_value(internal_element, 'Probability of impact')
    probability_of_impact_vulnerability = None
    if probability_of_impact_value is not None:
        probability_of_impact_vulnerability = probability_of_impact_value
    
    human_error_percentage_value = get_attribute_value(internal_element, 'HumanErrorEstimationPercentage')
    probability_of_human_error = None
    if human_error_percentage_value is not None:
        probability_of_human_error = calculate_probability_of_human_error(human_error_percentage_value, t)
    
    internal_element_data = {
        'ID': internal_element_id,
        'Name': internal_element_name,
        'Probability of failure': probability_of_failure,
        'Probability of Exposure': probability_of_exposure,
        'Probability of Impact' : probability_of_impact_vulnerability,
        'Probability of human error': probability_of_human_error,
        'RefBaseSystemUnitPath': ref_base_system_unit_path
    }
    
    # Categorize elements based on RefBaseSystemUnitPath
    if ref_base_system_unit_path in AutomationPyramidStatus:
        AutomationPyramidStatus[ref_base_system_unit_path].append(internal_element_data)
    elif ref_base_system_unit_path == 'HazardforSystem/Hazard':
        HazardinSystem.append(internal_element_data)
    elif ref_base_system_unit_path == 'VulnerabilityforSystem/Vulnerability':
        VulnerabilityinSystem.append(internal_element_data)
    
    # Add all probability data to a single list
    probability_data.append(internal_element_data)

# List to store the extracted information
external_interfaces_list = []
# Find all InternalElement elements
#internal_elements = root.findall('.//{http://www.dke.de/CAEX}InternalElement')

# Iterate over InternalElement elements
for internal_element in internal_elements:
    # Find all ExternalInterface elements within the current InternalElement
    #external_interfaces = internal_element.findall('.//{http://www.dke.de/CAEX}ExternalInterface')
    external_interfaces = internal_element.findall(externalInterfaceTag)

    
    # Check if the current InternalElement has ExternalInterface elements
    if len(external_interfaces) < 5: # there are 5 external interfaces links Network, process, User, hazard, and vulnerability based
        # Extract the attributes of the InternalElement
        internal_element_id = internal_element.get('ID')
        internal_element_name = internal_element.get('Name')

        # Iterate over ExternalInterface elements
        for external_interface in external_interfaces:
            # Extract the attributes of the ExternalInterface
            external_interface_id = external_interface.get('ID')
            external_interface_name = external_interface.get('Name')
            external_interface_ref_base_class_path = external_interface.get('RefBaseClassPath')
            if external_interface_ref_base_class_path != 'ConnectionBetnAssets/Network based':
            # Create a dictionary to store the extracted information
                external_interface_info = {
                    'InternalElement ID': internal_element_id,
                    'InternalElement Name': internal_element_name,
                    'ExternalInterface ID': external_interface_id,
                    'ExternalInterface Name': external_interface_name,
                    'ExternalInterface RefBaseClassPath': external_interface_ref_base_class_path
                    }
            
                # Add the dictionary to the list
                external_interfaces_list.append(external_interface_info)
                
        
# Rest of the code for extracting attributes and creating external_interfaces_list...

connections = []
interface_to_element_map = {}  # Dictionary to map external_interface_id to internal_element_id

for external_interface in external_interfaces_list:
    external_interface_id = external_interface['ExternalInterface ID']
    internal_element_id = external_interface['InternalElement ID']
    interface_to_element_map[external_interface_id] = internal_element_id

for internal_link in root.findall(internalLinkTag):
    ref_partner_a = internal_link.get('RefPartnerSideA')
    ref_partner_b = internal_link.get('RefPartnerSideB')

    if ref_partner_a in interface_to_element_map and ref_partner_b in interface_to_element_map:
        internal_element_a = interface_to_element_map[ref_partner_a]
        internal_element_b = interface_to_element_map[ref_partner_b]
        connection = {'from': internal_element_a, 'to': internal_element_b}
        connections.append(connection)

connections_mapped = []

for connection in connections:
    from_interface = connection['from']
    to_interface = connection['to']
    
    # Map the from_interface and to_interface to internal element IDs
    if from_interface in interface_to_element_map:
        from_element = interface_to_element_map[from_interface]
    else:
        from_element = from_interface  # If from_interface is not in the map, assume it's already an internal element ID
        
    if to_interface in interface_to_element_map:
        to_element = interface_to_element_map[to_interface]
    else:
        to_element = to_interface  # If to_interface is not in the map, assume it's already an internal element ID
    
    # Create the mapped connection
    mapped_connection = {'from': from_element, 'to': to_element}
    connections_mapped.append(mapped_connection)

# Create a defaultdict to store the connections
connections_from_to = defaultdict(list)
connections_to_from = defaultdict(list)


# Iterate over the connections and populate the connections_dict
total_elements = set() #number of nodes
for connection in connections_mapped:
    from_element = connection['from']
    to_element = connection['to']
    total_elements.add(from_element)
    total_elements.add(to_element)

    connections_from_to[from_element].append(to_element)
    connections_to_from[to_element].append(from_element)


# Convert the connections_dict to a list of dictionaries
connections_result_FT = [{'from': k, 'to': v} for k, v in connections_from_to.items()]
connections_result_TF = [{'from': v, 'to': k} for k, v in connections_to_from.items()]

number_of_children =  [{'Element': k, 'Number of children': len(v)} for k, v in connections_from_to.items()]
number_of_parents =  [{'Element': k, 'Number of parents': len(v)} for k, v in connections_to_from.items()]

"""To get the number of nodes connected to a particular node for impact analysis and priortisation"""

result_list = []
number_of_dependents = []
max_num_parents = 0
max_num_children = 0
for element in total_elements:
    # Assume number_of_children and number_of_parents are lists of dictionaries
    child = next((c for c in number_of_children if c['Element'] == element), {'Number of children': 0})
    parent = next((p for p in number_of_parents if p['Element'] == element), {'Number of parents': 0})
    
    total_dependents = child['Number of children'] + parent['Number of parents']
    
    # Create a new dictionary with the total dependents
    result_dict = {
        'Element': element,
        'Number of children': child['Number of children'],
        'Number of parents': parent['Number of parents'],
        'Total Dependents': total_dependents
    }

    for key in result_dict:
        if isinstance(result_dict[key], (int, float)):
            result_dict[key] /= len(total_elements)
    
    result_list.append(result_dict)
    parent = next((p for p in number_of_parents if p['Element'] == element), {'Number of parents': 0})
    
    num_parents = parent['Number of parents']
    if num_parents > max_num_parents:
        max_num_parents = num_parents
    
    num_children = child['Number of children']
    if num_children > max_num_children:
        max_num_children = num_children

#Bayesian Program Starts from Here
'''%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%'''

# Create an empty Bayesian Model
bbn = BayesianNetwork()

# Define the nodes in the BBN
connections = connections_mapped

# Add nodes to the BBN
bbnNodes=bbn.add_nodes_from(total_elements)

# Add edges to the BBN
bbn.add_edges_from([(connection['from'], connection['to']) for connection in connections])
# Define the conditional probability distributions (CPDs) for each node
cpds = {}
cpd_values_list = [] # to store the CPD values
nodes_and_numberofParents =[]
sap_percent = float(input('What is the probability of a successful security attack? \n (0.01 < value < 10))\n'))
if 0.01 <= sap_percent <= 10:
    sap = sap_percent / 100
else:
    print('Wrong Value: Contact security personnel')
def generate_cpd_values_occur(num_states, num_parents, hazard_node=False, vulnerability_node=False, process_node=False):
    cpd_values = np.zeros((num_states, 2 ** num_parents))
    #random_number = random.uniform(0.1, 0.5)
    
    if hazard_node:
        if num_parents == 0:
            cpd_values[0, 0] = 0.5
            cpd_values[1, 0] = 0.5
        elif num_parents == 1:
            cpd_values[0, 0] = 1
            cpd_values[0, 1] = 0
            cpd_values[1, 0] = 1 - cpd_values[0, 0]
            cpd_values[1, 1] = 1 - cpd_values[0, 1]
            
        elif 2<=num_parents <= max_num_parents:
            cpd_values=generate_cpd_values_hazard(num_parents)
    
    elif vulnerability_node:
        # Generate CPD values for vulnerability nodes
        probability_of_exposure_for_node = matching_vulnerability_nodes[0]['Probability of Exposure']
        pofe = float(probability_of_exposure_for_node)
        
        if num_parents == 0:
            cpd_values[0, 0] = pofe*sap
            cpd_values[1, 0] = 1 - pofe*sap
        elif num_parents >= 1:
            cpd_values[0, :-1] = pofe*sap  	# parent Vulnerability is exposed
            cpd_values[1, :-1] = 1-pofe*sap # parent vulnerability is unexposed
            cpd_values[0, -1] = 0	# parent vulnerability is exposed
            cpd_values[1, -1] = 1
            cpd_values[0, 0] = 1
            cpd_values[1, 0] = 0
    elif process_node:
        # Generate CPD values for process nodes
        ref_base_for_node = matching_process_nodes[0]['RefBaseSystemUnitPath']
        
        if ref_base_for_node in ['AssetOfICS/Hardware/Process device/Actuator',
                                 'AssetOfICS/Hardware/Process device/Sensor',
                                 'AssetOfICS/Hardware/Machine', 
                                 'AssetOfICS/Hardware/Process device/Controller',
                                 'AssetOfICS/Hardware/Process device/Work station']:
            probability_of_failure_for_node = matching_process_nodes[0]['Probability of failure']
            if probability_of_failure_for_node:
                poff = float(probability_of_failure_for_node)
                cpd_values[0, :-1] = 1.0
                cpd_values[1, :-1] = 0.0
                cpd_values[0, -1] = poff
                cpd_values[1, -1] = 1 - poff
            else:
                cpd_values[0, :-1] = 1.0
                cpd_values[1, :-1] = 0.0
                cpd_values[0, -1] = 0.0
                cpd_values[1, -1] = 1
                
        elif ref_base_for_node == 'AssetOfICS/User':
            probability_of_human_error_for_node = matching_process_nodes[0]['Probability of human error']
            pofhe = float(probability_of_human_error_for_node)
            cpd_values[0, 0] = pofhe
            cpd_values[1, 0] = 1 - pofhe

        else:
            probability_of_failure_for_node = matching_process_nodes[0]['Probability of failure']
            poff = float(probability_of_failure_for_node)
            cpd_values[0, 0] = poff
            cpd_values[1, 0] = 1 - poff

    cpd_values /= np.sum(cpd_values, axis=0)  # Normalize the CPD values

    return cpd_values.reshape((num_states, -1))

def generate_cpd_values_impact_(num_states, num_parents, hazard_node=False, vulnerability_node=False, process_node=False):
    cpd_values = np.zeros((num_states, 2 ** num_parents))
    #random_number = random.uniform(0.1, 0.5)
    current_entry = next((entry for entry in result_list if entry['Element'] == node), None)
    if hazard_node:
        if num_parents == 0:
            cpd_values[0, 0] = 0.5
            cpd_values[1, 0] = 0.5
        elif num_parents == 1:
            cpd_values[0, 0] = 1
            cpd_values[0, 1] = 0
            cpd_values[1, 0] = 1 - cpd_values[0, 0]
            cpd_values[1, 1] = 1 - cpd_values[0, 1]
            
        elif 2<=num_parents <= max_num_parents:
            cpd_values=generate_cpd_values_impact(num_parents)
    
    elif vulnerability_node:
        # Generate CPD values for vulnerability nodes
        probability_of_exposure_for_node = matching_vulnerability_nodes[0]['Probability of Impact']
        pofe = float(probability_of_exposure_for_node)
        
        if num_parents == 0:
            cpd_values[0, 0] = pofe*sap
            cpd_values[1, 0] = 1 - pofe*sap
        elif num_parents >= 1:
            cpd_values[0, :-1] = 1  	# parent Vulnerability is exposed
            cpd_values[1, :-1] = 0 # parent vulnerability is unexposed
            cpd_values[0, -1] = pofe*sap	# parent vulnerability is exposed
            cpd_values[1, -1] = 1 - pofe*sap	# parent vulnerability is unexposed
    elif process_node:
        # Generate CPD values for process nodes
        ref_base_for_node = matching_process_nodes[0]['RefBaseSystemUnitPath']
        
        if ref_base_for_node in ['AssetOfICS/Hardware/Process device/Actuator',
                                 'AssetOfICS/Hardware/Process device/Sensor',
                                 'AssetOfICS/Hardware/Machine', 
                                 'AssetOfICS/Hardware/Process device/Controller',
                                 'AssetOfICS/Hardware/Process device/Work station']:
            probability_of_failure_for_node = matching_process_nodes[0]['Probability of failure']
            if probability_of_failure_for_node:
                cpd_values[0, :-1] = 1.0
                cpd_values[1, :-1] = 0.0
                cpd_values[0, -1] = current_entry['Number of children']
                cpd_values[1, -1] = 1 - current_entry['Number of children']
            else:
                cpd_values[0, :-1] = 1.0
                cpd_values[1, :-1] = 0.0
                cpd_values[0, -1] = 0.0
                cpd_values[1, -1] = 1
                
        elif ref_base_for_node == 'AssetOfICS/User':
            cpd_values[0, 0] = current_entry['Number of children']
            cpd_values[1, 0] = 1 - current_entry['Number of children']

        else:
            cpd_values[0, 0] = current_entry['Number of children']
            cpd_values[1, 0] = 1 - current_entry['Number of children']

    cpd_values /= np.sum(cpd_values, axis=0)  # Normalize the CPD values

    return cpd_values.reshape((num_states, -1))

cpd_values_list = []
# Define the CPD values for each node based on your specific requirements
for node in bbn.nodes():
    num_parents = len(bbn.get_parents(node))
    num_states = 2  # Assuming binary states for each node

    matching_hazard_nodes = [element for element in HazardinSystem if element['ID'] == node]
    matching_vulnerability_nodes = [element for element in VulnerabilityinSystem if element['ID'] == node]
    matching_process_nodes = [element for element in probability_data if element['ID'] == node]

    cpd_values = None

    if matching_hazard_nodes:
        hazard_node = True
        cpd_values = generate_cpd_values_occur(num_states, num_parents, hazard_node=True)
    elif matching_vulnerability_nodes:
        vulnerability_node = True
        cpd_values = generate_cpd_values_occur(num_states, num_parents, vulnerability_node=True)
    elif matching_process_nodes:
        process_node = True
        cpd_values = generate_cpd_values_occur(num_states, num_parents, process_node=True)

    # Create the TabularCPD object for the node
    cpd = TabularCPD(variable=node, variable_card=num_states, values=cpd_values,
                     evidence=bbn.get_parents(node), evidence_card=[2] * num_parents)

    cpds[node] = cpd
    cpd_values_list.append((node, cpd_values.tolist(), cpd.variables, cpd.cardinality))
    
# Add CPDs to the BBN
bbn.add_cpds(*cpds.values())
import itertools
bbn_graph = bbn.to_markov_model()

# Find the last node in the graph
last_node = None
for element1 in HazardinSystem:
    node1=element1['ID']
    for element2 in result_list:
        node2=element2['Element']
        child_num = element2['Number of children']
        if node1 == node2:
            if child_num == 0.0:
                last_node = node1

print("Last node in the BBN graph:", last_node)

# Function to calculate the shortest path length between two nodes
def shortest_path_length(graph, start_node, end_node):
    try:
        # Using networkx's shortest path length function
        length = nx.shortest_path_length(graph, source=start_node, target=end_node)
        return length
    except nx.NetworkXNoPath:
        # If no path exists between the nodes
        return float('inf')

# Function to randomly select start and end nodes from total_elements
def select_start_end_nodes(total_elements):
    start_node = random.choice(list(total_elements))
    end_node = random.choice(list(total_elements - {start_node}))  # Ensuring end_node is different from start_node
    return start_node, end_node

path_length_betn_nodes=[]
path_length_betn_nodes_final=[]
path_length_final_node = []
# Loop through each element in total_elements
for node1, node2 in itertools.product(total_elements, repeat=2):
    if node1 == node2:
        #print(f"Number of hops between {node1} and {node2}: 0")
        path_length_betn_nodes.append((node1, node2, 0))
    else:
        path_length = shortest_path_length(bbn_graph, node1, node2)
        if path_length == float('inf'):
            #print(f"There is no path between {node1} and {node2}.")
            path_length_betn_nodes.append((node1, node2, "No path"))
        else:
            #print(f"Number of hops between {node1} and {node2}: {path_length}")
            path_length_betn_nodes_final.append((node1, node2, path_length, 1/path_length))
            path_length_betn_nodes.append({'Node1': node1, 'Node2': node2, 
                               'Number of hops': path_length, 
                               'Probability': 1/path_length})
            if node2==last_node:
                path_length_final_node.append((node1, last_node, path_length, 1/path_length))

''''Impact Program'''
bbn_impact=BayesianNetwork()

bbn_impact.add_edges_from([(connection['from'], connection['to']) for connection in connections])

cpds = {}
cpd_values_list_impact = [] # to store the CPD values
nodes_and_numberofParents =[]

for node in bbn_impact.nodes():
    num_parents = len(bbn.get_parents(node))
    num_states = 2  # Assuming binary states for each node

    matching_hazard_nodes = [element for element in HazardinSystem if element['ID'] == node]
    matching_vulnerability_nodes = [element for element in VulnerabilityinSystem if element['ID'] == node]
    matching_process_nodes = [element for element in probability_data if element['ID'] == node]

    cpd_values = None

    if matching_hazard_nodes:
        hazard_node = True
        cpd_values = generate_cpd_values_impact_(num_states, num_parents, hazard_node=True)
    elif matching_vulnerability_nodes:
        vulnerability_node = True
        cpd_values = generate_cpd_values_impact_(num_states, num_parents, vulnerability_node=True)
    elif matching_process_nodes:
        process_node = True
        cpd_values = generate_cpd_values_impact_(num_states, num_parents, process_node=True)

    # Create the TabularCPD object for the node
    cpd = TabularCPD(variable=node, variable_card=num_states, values=cpd_values,
                     evidence=bbn.get_parents(node), evidence_card=[2] * num_parents)

    cpds[node] = cpd
    cpd_values_list.append((node, cpd_values.tolist(), cpd.variables, cpd.cardinality))

# Add CPDs to the Bayesian Model
bbn_impact.add_cpds(*cpds.values())

''''Impact program ends'''

# Check if the BBN structure and CPDs are consistent
print("BBN structure for occurence frequency is consistent:", bbn.check_model())
print("BBN structure for severity is consistent:", bbn_impact.check_model())

inference = VariableElimination(bbn)
inference2 = VariableElimination(bbn_impact)

# Create a directed graph
graph = nx.DiGraph()

# Add nodes to the graph
graph.add_nodes_from(bbn.nodes())

# Add edges to the graph
graph.add_edges_from(bbn.edges())

# Set the layout of the graph
#pos = nx.drawing.layout.shell_layout(graph, scale=2) #Shell layout places all the nodes in a circle and edges are in between
pos = nx.kamada_kawai_layout(graph, scale=2)         #kamada layout places all the nodes in a parent child fashion
#pos = nx.spectral_layout(graph)             #spectral layout places all the nodes in a parent child fashion, but not properly as kamada
#pos = nx.spiral_layout(graph, scale=5)             

# Draw the nodes and edges
nx.draw_networkx_nodes(graph, pos, node_color='lightblue', node_size=300)
nx.draw_networkx_edges(graph, pos, arrows=True, arrowstyle='->', arrowsize=10)
nx.draw_networkx_labels(graph, pos)

# Adjust plot settings
plt.title("Bayesian Belief Network")
plt.axis('off')

# Save the plot to a file
plt.savefig('bbn_plot_shell_layout.png', format='png', dpi=300, bbox_inches='tight')

# Show the plot
#plt.show()
risk_measurements = []
for node in total_elements:
    prob_node = inference.query(variables=[node])
    impact_node = inference2.query(variables=[node])
    for element in path_length_final_node:
        if node == element[0]:
            cpd_prob = prob_node.values
            cpd_imp = element[3]
            cpd_impact = impact_node.values        
            risk_v1=cpd_prob[0] * cpd_imp
            risk_v2=cpd_prob[0] * cpd_impact[0]
            #print(cpd_prob[0], cpd_imp, risk)
            risk_measurements.append([node, cpd_prob[0], cpd_imp, cpd_impact[0], risk_v1, risk_v2])

import openpyxl
from openpyxl.styles import PatternFill

# Initialize an empty dictionary to store node and corresponding normal_prob_P3L_if_node_fail values
node_prob_dict = {}

# Loop through the elements
for node in total_elements:
    if node == last_node:
        pass
    else:
        prob_P3L_if_node_fail = inference.query(variables=[last_node], evidence={node:0})
        prob_P3L_if_node_fail_values = prob_P3L_if_node_fail.values
        
        min_value = min(prob_P3L_if_node_fail_values)
        max_value = max(prob_P3L_if_node_fail_values)
        
        min_value = round(min_value, 2) -0.01  # Restrict minimum value to 2 decimal places
        max_value = round(max_value, 2) + 0.01  # Restrict maximum value to 2 decimal places and add 0.01
        
        normal_prob_P3L_if_node_fail = (prob_P3L_if_node_fail_values[0] - min_value) / (max_value - min_value)
        
        # Store node and corresponding normal_prob_P3L_if_node_fail values in the dictionary
        node_prob_dict[node] = normal_prob_P3L_if_node_fail

# Sort the dictionary by values in descending order
sorted_node_prob = sorted(node_prob_dict.items(), key=lambda x: x[1], reverse=True)

# Create an Excel workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers
worksheet['A1'] = 'Node'
worksheet['B1'] = 'Probability'

# Define colors
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Write sorted results to Excel file
for index, (node, prob) in enumerate(sorted_node_prob, start=2):
    cell_node = worksheet.cell(row=index, column=1)
    cell_node.value = node
    
    cell_prob = worksheet.cell(row=index, column=2)
    cell_prob.value = prob
    
    if prob > 0.8:
        cell_prob.fill = red_fill
    elif prob >= 0.5 and prob <= 0.8:
        cell_prob.fill = yellow_fill
    else:
        cell_prob.fill = green_fill

# Save the workbook
workbook.save(filename='sorted_results.xlsx')


if user_choice==1:
    for nodes in total_elements:
        if nodes==last_node:
            prob_last_node=inference.query(variables=[nodes])
            impact_P3L=inference2.query(variables=[nodes])
            cpd_prob = prob_last_node.values
            cpd_impact = impact_P3L.values
            #print(cpd_prob, cpd_impact)
            risk_free_prob = cpd_prob[1]*cpd_impact[1]
            risk_prob = 1-risk_free_prob
            #print(risk_prob)
            #normal_risk = (risk_prob - 0.64631985)/(0.665071318 - 0.64631985)
            normal_risk = (risk_prob - 0.22495)/(0.58217 - 0.22495)
            #print(normal_risk)
            
            # Normalize the probabilities
            print('The current risk is : {:.2f} %'.format(normal_risk * 100))
            if normal_risk<0.2:
                #print('The current risk of failure due to safety or security is : {:.2f} %'.format(cpd[0] * 100))
                print('The system is under NEGLIGIBLE risk')
            elif 0.2<=normal_risk<0.4:
                #print('The current risk of failure due to safety or security is : {:.2f} %'.format(cpd[0] * 100))
                print('The system is under LOW risk')
            elif 0.4<=normal_risk<0.6:
                #print('The current risk of failure due to safety or security is : {:.2f} %'.format(cpd[0] * 100))
                print('The system is under MEDIUM risk')
            elif 0.6<=normal_risk<0.8:
                #print('The current risk of failure due to safety or security is : {:.2f} %'.format(cpd[0] * 100))
                print('The system is under HIGH risk')
            else:
                #print('The current risk of failure due to safety or security is : {:.2f} %'.format(cpd[0] * 100))
                print('The system is under CRITICAL risk')
        else:
            pass
elif user_choice==2:
    for nodes in total_elements:
        if nodes==last_node:
            prob_SPI=inference.query(variables=[nodes])
            impact_SPI=inference2.query(variables=[nodes])
            print(prob_SPI)
            #print(impact_SPI)
            prob_of_process_safety=inference.query(variables=[nodes], evidence={'V1':1, 'V2':1, 'V3':1, 'V4':1, 'V5':1})
            print(prob_of_process_safety)
            
            cpd_prob = prob_SPI.values
            cpd_impact = impact_SPI.values
            print(cpd_prob)
            #print(cpd_impact)
            risk_prob = cpd_prob[0]*cpd_impact[0]
            #risk_prob = 1-risk_free_prob
            print(risk_prob)
            normal_risk = (risk_prob)/(0.0674)
            print(normal_risk)
            
            # Normalize the probabilities
            print('The current risk is : {:.2f} %'.format(normal_risk * 100))
            if normal_risk<0.2:
                #print('The current risk of failure due to safety or security is : {:.2f} %'.format(cpd[0] * 100))
                print('The system is under NEGLIGIBLE risk')
            elif 0.2<=normal_risk<0.4:
                #print('The current risk of failure due to safety or security is : {:.2f} %'.format(cpd[0] * 100))
                print('The system is under LOW risk')
            elif 0.4<=normal_risk<0.6:
                #print('The current risk of failure due to safety or security is : {:.2f} %'.format(cpd[0] * 100))
                print('The system is under MEDIUM risk')
            elif 0.6<=normal_risk<0.8:
                #print('The current risk of failure due to safety or security is : {:.2f} %'.format(cpd[0] * 100))
                print('The system is under HIGH risk')
            else:
                #print('The current risk of failure due to safety or security is : {:.2f} %'.format(cpd[0] * 100))
                print('The system is under CRITICAL risk')
        else:
            pass
        
else:
    print('Select the right number')
